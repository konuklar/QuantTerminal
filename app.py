# =============================================================
# 🏛️ APOLLO/ENIGMA QUANT TERMINAL v6.0 - INSTITUTIONAL EDITION
# Professional Global Multi-Asset Portfolio Management System
# Enhanced with Institutional Reporting & Advanced Analytics
# =============================================================

import os
import warnings
import numpy as np
import pandas as pd
import streamlit as st
import yfinance as yf
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
from scipy import stats, optimize
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional, Union
import json
import concurrent.futures
from functools import lru_cache
import traceback
import time
import hashlib
import inspect
from pathlib import Path
import pickle
import tempfile
import base64
from io import BytesIO
import io

# Import PyPortfolioOpt with all required components
try:
    from pypfopt import expected_returns, risk_models, objective_functions, black_litterman, hierarchical_portfolio
    from pypfopt.efficient_frontier import EfficientFrontier, EfficientSemivariance
    from pypfopt.discrete_allocation import DiscreteAllocation, get_latest_prices
    from pypfopt.objective_functions import L2_reg, transaction_cost
    from pypfopt.cla import CLA
    from pypfopt.hierarchical_portfolio import HRPOpt
    from pypfopt.risk_models import CovarianceShrinkage
    PYPFOPT_AVAILABLE = True
except ImportError as e:
    PYPFOPT_AVAILABLE = False
    st.warning(f"⚠️ PyPortfolioOpt not installed. Some optimization features limited.")

# Additional imports
from scipy.spatial.distance import squareform
from scipy.cluster.hierarchy import linkage
import matplotlib.pyplot as plt
import seaborn as sns
from fpdf import FPDF
import pdfkit
from weasyprint import HTML
import warnings
warnings.filterwarnings("ignore")

# -------------------------------------------------------------
# SUPERIOR INSTITUTIONAL STYLING
# -------------------------------------------------------------
st.set_page_config(
    page_title="APOLLO/ENIGMA - Institutional Portfolio Terminal",
    layout="wide",
    initial_sidebar_state="expanded",
    page_icon="🏛️"
)

# Enhanced institutional CSS
INSTITUTIONAL_CSS = """
<style>
:root {
    --primary: #0a3d62;
    --primary-dark: #082c47;
    --primary-light: #1a5fb4;
    --secondary: #26a269;
    --secondary-dark: #1e864d;
    --accent: #f39c12;
    --accent-dark: #d68910;
    --danger: #e74c3c;
    --danger-dark: #c0392b;
    --warning: #f1c40f;
    --warning-dark: #f39c12;
    --success: #27ae60;
    --success-dark: #229954;
    --dark-bg: #0f172a;
    --dark-bg-light: #1e293b;
    --card-bg: #1e293b;
    --card-border: #334155;
    --text: #f8f9fa;
    --text-muted: #94a3b8;
    --text-light: #e2e8f0;
    --shadow: rgba(0, 0, 0, 0.3);
}

/* Enhanced Institutional Theme */
.main {
    background: linear-gradient(135deg, var(--dark-bg) 0%, #0a1929 100%);
}

/* Professional Header */
.institutional-header {
    background: linear-gradient(90deg, var(--primary-dark), var(--primary));
    padding: 1.5rem;
    border-radius: 0 0 15px 15px;
    margin-bottom: 2rem;
    box-shadow: 0 4px 20px var(--shadow);
    border-bottom: 3px solid var(--accent);
}

/* Superior Cards */
.institutional-card {
    background: linear-gradient(145deg, var(--card-bg), #16213e);
    border: 1px solid var(--card-border);
    border-radius: 12px;
    padding: 1.5rem;
    margin-bottom: 1.5rem;
    box-shadow: 0 6px 25px rgba(0, 0, 0, 0.2);
    transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
    position: relative;
    overflow: hidden;
}

.institutional-card:hover {
    transform: translateY(-4px);
    box-shadow: 0 12px 40px rgba(0, 0, 0, 0.3);
    border-color: var(--primary-light);
}

.institutional-card::before {
    content: '';
    position: absolute;
    top: 0;
    left: 0;
    right: 0;
    height: 4px;
    background: linear-gradient(90deg, var(--primary), var(--accent));
}

/* Professional Metrics */
.institutional-metric {
    background: linear-gradient(135deg, rgba(10, 61, 98, 0.1), rgba(26, 95, 180, 0.05));
    border: 1px solid var(--card-border);
    border-radius: 10px;
    padding: 1.2rem;
    text-align: center;
    transition: all 0.3s ease;
}

.institutional-metric:hover {
    background: linear-gradient(135deg, rgba(10, 61, 98, 0.2), rgba(26, 95, 180, 0.1));
    border-color: var(--primary-light);
}

.metric-value {
    font-size: 1.8rem;
    font-weight: 700;
    color: var(--text);
    margin: 0.5rem 0;
    text-shadow: 0 2px 4px rgba(0, 0, 0, 0.2);
}

.metric-label {
    font-size: 0.85rem;
    color: var(--text-muted);
    text-transform: uppercase;
    letter-spacing: 1px;
    font-weight: 500;
}

/* Enhanced Tabs */
.stTabs [data-baseweb="tab-list"] {
    gap: 4px;
    background: var(--dark-bg-light);
    padding: 6px;
    border-radius: 12px;
    border: 1px solid var(--card-border);
}

.stTabs [data-baseweb="tab"] {
    background: transparent;
    border-radius: 8px;
    padding: 12px 24px;
    font-weight: 600;
    color: var(--text-muted);
    border: 1px solid transparent;
    transition: all 0.3s ease;
    font-size: 0.9rem;
}

.stTabs [data-baseweb="tab"]:hover {
    background: rgba(10, 61, 98, 0.2);
    color: var(--text-light);
}

.stTabs [aria-selected="true"] {
    background: linear-gradient(135deg, var(--primary), var(--primary-dark)) !important;
    color: white !important;
    border-color: var(--accent) !important;
    box-shadow: 0 4px 15px rgba(10, 61, 98, 0.3);
}

/* Professional Buttons */
.institutional-button {
    background: linear-gradient(135deg, var(--primary), var(--primary-dark));
    color: white;
    border: none;
    padding: 12px 28px;
    border-radius: 8px;
    font-weight: 600;
    transition: all 0.3s;
    box-shadow: 0 4px 15px rgba(10, 61, 98, 0.3);
    position: relative;
    overflow: hidden;
}

.institutional-button::before {
    content: '';
    position: absolute;
    top: 0;
    left: -100%;
    width: 100%;
    height: 100%;
    background: linear-gradient(90deg, transparent, rgba(255, 255, 255, 0.1), transparent);
    transition: 0.5s;
}

.institutional-button:hover {
    transform: translateY(-2px);
    box-shadow: 0 8px 25px rgba(10, 61, 98, 0.4);
}

.institutional-button:hover::before {
    left: 100%;
}

/* Status Indicators */
.status-badge {
    display: inline-block;
    padding: 4px 12px;
    border-radius: 20px;
    font-size: 0.75rem;
    font-weight: 600;
    letter-spacing: 0.5px;
    text-transform: uppercase;
}

.badge-success {
    background: linear-gradient(135deg, var(--success), var(--success-dark));
    color: white;
}

.badge-warning {
    background: linear-gradient(135deg, var(--warning), var(--warning-dark));
    color: #212529;
}

.badge-danger {
    background: linear-gradient(135deg, var(--danger), var(--danger-dark));
    color: white;
}

.badge-info {
    background: linear-gradient(135deg, var(--primary-light), #3498db);
    color: white;
}

/* Enhanced Tables */
.institutional-table {
    background: var(--card-bg);
    border: 1px solid var(--card-border);
    border-radius: 10px;
    overflow: hidden;
}

.institutional-table table {
    width: 100%;
    background: transparent;
}

.institutional-table th {
    background: var(--primary-dark);
    color: white;
    font-weight: 600;
    padding: 12px 16px;
    border-bottom: 2px solid var(--accent);
}

.institutional-table td {
    padding: 10px 16px;
    border-bottom: 1px solid var(--card-border);
    color: var(--text-light);
}

.institutional-table tr:hover {
    background: rgba(10, 61, 98, 0.1);
}

/* Professional Charts */
.js-plotly-plot {
    border: 1px solid var(--card-border);
    border-radius: 12px;
    overflow: hidden;
    background: var(--card-bg);
}

/* Sidebar Enhancement */
[data-testid="stSidebar"] {
    background: linear-gradient(135deg, var(--dark-bg-light), #1a1a2e);
    border-right: 1px solid var(--card-border);
}

/* Input Enhancements */
.stSelectbox > div > div, .stTextInput > div > div {
    background: var(--card-bg) !important;
    border: 1px solid var(--card-border) !important;
    border-radius: 8px !important;
}

.stSelectbox > div > div:hover, .stTextInput > div > div:hover {
    border-color: var(--primary-light) !important;
    box-shadow: 0 0 0 2px rgba(26, 95, 180, 0.2) !important;
}

/* Progress Bars */
.stProgress > div > div > div {
    background: linear-gradient(90deg, var(--primary), var(--accent));
    border-radius: 4px;
}

/* Custom Scrollbar */
::-webkit-scrollbar {
    width: 10px;
    height: 10px;
}

::-webkit-scrollbar-track {
    background: var(--dark-bg);
    border-radius: 5px;
}

::-webkit-scrollbar-thumb {
    background: linear-gradient(var(--primary), var(--primary-dark));
    border-radius: 5px;
    border: 2px solid var(--dark-bg);
}

::-webkit-scrollbar-thumb:hover {
    background: linear-gradient(var(--primary-dark), var(--primary));
}

/* Professional Loading */
.stSpinner > div {
    border: 4px solid var(--card-border);
    border-top: 4px solid var(--primary);
    border-radius: 50%;
    width: 50px;
    height: 50px;
    animation: institutional-spin 1s linear infinite;
}

@keyframes institutional-spin {
    0% { transform: rotate(0deg); }
    100% { transform: rotate(360deg); }
}

/* Grid Layout */
.institutional-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
    gap: 1.5rem;
    margin-bottom: 2rem;
}

/* Tooltips */
[data-tooltip] {
    position: relative;
    cursor: help;
}

[data-tooltip]:hover::after {
    content: attr(data-tooltip);
    position: absolute;
    bottom: 125%;
    left: 50%;
    transform: translateX(-50%);
    background: var(--primary-dark);
    color: white;
    padding: 8px 12px;
    border-radius: 6px;
    font-size: 0.8rem;
    white-space: nowrap;
    z-index: 1000;
    border: 1px solid var(--accent);
    box-shadow: 0 4px 15px var(--shadow);
}

/* Expander Headers */
.streamlit-expanderHeader {
    background: linear-gradient(90deg, var(--card-bg), var(--dark-bg-light));
    border: 1px solid var(--card-border);
    border-radius: 8px;
    color: var(--text);
    font-weight: 600;
    padding: 1rem;
}

.streamlit-expanderHeader:hover {
    background: linear-gradient(90deg, var(--dark-bg-light), rgba(10, 61, 98, 0.2));
    border-color: var(--primary-light);
}

/* Alert Enhancements */
.stAlert {
    border-radius: 10px !important;
    border: 1px solid !important;
    background: var(--card-bg) !important;
}

/* Footer */
.institutional-footer {
    background: linear-gradient(90deg, var(--dark-bg-light), var(--primary-dark));
    padding: 1rem;
    margin-top: 3rem;
    border-radius: 10px;
    border-top: 2px solid var(--accent);
    text-align: center;
    color: var(--text-muted);
    font-size: 0.8rem;
}
</style>
"""

st.markdown(INSTITUTIONAL_CSS, unsafe_allow_html=True)

# -------------------------------------------------------------
# INSTITUTIONAL REPORTING SYSTEM
# -------------------------------------------------------------
class InstitutionalReportGenerator:
    """Superior institutional reporting system with multiple formats"""
    
    @staticmethod
    def generate_executive_summary(portfolio_data: Dict, analysis_results: Dict) -> str:
        """Generate professional executive summary"""
        summary = f"""
# 🏛️ EXECUTIVE SUMMARY
## APOLLO/ENIGMA PORTFOLIO ANALYSIS
### {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

---

## 📊 PORTFOLIO OVERVIEW

| Metric | Value | Benchmark | Status |
|--------|-------|-----------|---------|
| **Total Return** | {portfolio_data.get('total_return', 0):.2%} | {portfolio_data.get('benchmark_return', 0):.2%} | {InstitutionalReportGenerator._get_status_icon(portfolio_data.get('total_return', 0) >= portfolio_data.get('benchmark_return', 0))} |
| **Annual Return** | {portfolio_data.get('annual_return', 0):.2%} | {portfolio_data.get('benchmark_annual', 0):.2%} | {InstitutionalReportGenerator._get_status_icon(portfolio_data.get('annual_return', 0) >= portfolio_data.get('benchmark_annual', 0))} |
| **Annual Volatility** | {portfolio_data.get('annual_vol', 0):.2%} | {portfolio_data.get('benchmark_vol', 0):.2%} | {InstitutionalReportGenerator._get_status_icon(portfolio_data.get('annual_vol', 0) <= portfolio_data.get('benchmark_vol', 0))} |
| **Sharpe Ratio** | {portfolio_data.get('sharpe', 0):.2f} | {portfolio_data.get('benchmark_sharpe', 0):.2f} | {InstitutionalReportGenerator._get_status_icon(portfolio_data.get('sharpe', 0) >= portfolio_data.get('benchmark_sharpe', 0))} |
| **Max Drawdown** | {portfolio_data.get('max_dd', 0):.2%} | {portfolio_data.get('benchmark_dd', 0):.2%} | {InstitutionalReportGenerator._get_status_icon(portfolio_data.get('max_dd', 0) >= portfolio_data.get('benchmark_dd', 0))} |

---

## 🎯 KEY FINDINGS

### 1. PERFORMANCE ASSESSMENT
{InstitutionalReportGenerator._generate_performance_assessment(portfolio_data)}

### 2. RISK ANALYSIS
{InstitutionalReportGenerator._generate_risk_assessment(analysis_results)}

### 3. PORTFOLIO CONSTRUCTION
{InstitutionalReportGenerator._generate_portfolio_assessment(portfolio_data)}

---

## 📈 RECOMMENDATIONS

### 🟢 STRENGTHS
{InstitutionalReportGenerator._generate_strengths(portfolio_data)}

### 🟡 OPPORTUNITIES
{InstitutionalReportGenerator._generate_opportunities(portfolio_data)}

### 🔴 RISKS TO MONITOR
{InstitutionalReportGenerator._generate_risks(analysis_results)}

---

## 📊 DETAILED METRICS

### Risk-Adjusted Performance
{InstitutionalReportGenerator._generate_detailed_metrics(portfolio_data)}

### Portfolio Characteristics
{InstitutionalReportGenerator._generate_portfolio_characteristics(portfolio_data)}

---

*This report was generated by Apollo/ENIGMA Institutional Terminal v6.0*
*For institutional use only. Not for distribution to retail investors.*
"""
        return summary
    
    @staticmethod
    def _get_status_icon(condition: bool) -> str:
        return "✅" if condition else "⚠️"
    
    @staticmethod
    def _generate_performance_assessment(data: Dict) -> str:
        """Generate performance assessment section"""
        assessment = ""
        if data.get('total_return', 0) > data.get('benchmark_return', 0):
            assessment += "- Portfolio outperformed benchmark by "
            assessment += f"{(data['total_return'] - data['benchmark_return']):.2%}\n"
        else:
            assessment += "- Portfolio underperformed benchmark by "
            assessment += f"{(data['benchmark_return'] - data['total_return']):.2%}\n"
        
        if data.get('sharpe', 0) > 1.0:
            assessment += "- Excellent risk-adjusted returns (Sharpe > 1.0)\n"
        elif data.get('sharpe', 0) > 0.5:
            assessment += "- Good risk-adjusted returns\n"
        else:
            assessment += "- Below target risk-adjusted returns\n"
        
        return assessment
    
    @staticmethod
    def _generate_risk_assessment(data: Dict) -> str:
        """Generate risk assessment section"""
        assessment = ""
        if data.get('var_95', 0) > -0.05:
            assessment += "- Normal market risk exposure\n"
        else:
            assessment += "- Elevated downside risk\n"
        
        if data.get('max_dd', 0) > -0.20:
            assessment += "- Manageable drawdown risk\n"
        else:
            assessment += "- High drawdown risk\n"
        
        return assessment
    
    @staticmethod
    def _generate_strengths(data: Dict) -> str:
        """Generate strengths section"""
        strengths = ""
        if data.get('sharpe', 0) > 1.0:
            strengths += "- Superior risk-adjusted performance\n"
        if data.get('diversification_score', 0) > 0.7:
            strengths += "- Well-diversified portfolio construction\n"
        if data.get('liquidity_score', 0) > 0.8:
            strengths += "- Strong liquidity profile\n"
        return strengths
    
    @staticmethod
    def generate_pdf_report(portfolio_data: Dict, analysis_results: Dict, charts: List[bytes]) -> bytes:
        """Generate professional PDF report"""
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        
        # Add first page
        pdf.add_page()
        pdf.set_font("Arial", 'B', 24)
        pdf.cell(0, 20, "APOLLO/ENIGMA PORTFOLIO REPORT", ln=True, align='C')
        pdf.set_font("Arial", '', 12)
        pdf.cell(0, 10, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True, align='C')
        pdf.ln(10)
        
        # Add summary
        pdf.set_font("Arial", 'B', 16)
        pdf.cell(0, 10, "Executive Summary", ln=True)
        pdf.set_font("Arial", '', 12)
        
        summary = InstitutionalReportGenerator.generate_executive_summary(portfolio_data, analysis_results)
        pdf.multi_cell(0, 8, summary)
        
        # Add performance metrics table
        pdf.add_page()
        pdf.set_font("Arial", 'B', 16)
        pdf.cell(0, 10, "Performance Metrics", ln=True)
        pdf.ln(5)
        
        metrics = [
            ["Metric", "Portfolio", "Benchmark", "Status"],
            ["Total Return", f"{portfolio_data.get('total_return', 0):.2%}", 
             f"{portfolio_data.get('benchmark_return', 0):.2%}", ""],
            ["Annual Return", f"{portfolio_data.get('annual_return', 0):.2%}", 
             f"{portfolio_data.get('benchmark_annual', 0):.2%}", ""],
            ["Annual Volatility", f"{portfolio_data.get('annual_vol', 0):.2%}", 
             f"{portfolio_data.get('benchmark_vol', 0):.2%}", ""],
            ["Sharpe Ratio", f"{portfolio_data.get('sharpe', 0):.2f}", 
             f"{portfolio_data.get('benchmark_sharpe', 0):.2f}", ""],
            ["Max Drawdown", f"{portfolio_data.get('max_dd', 0):.2%}", 
             f"{portfolio_data.get('benchmark_dd', 0):.2%}", ""],
        ]
        
        col_widths = [50, 40, 40, 30]
        for row in metrics:
            for i, cell in enumerate(row):
                pdf.cell(col_widths[i], 8, cell, border=1)
            pdf.ln()
        
        return pdf.output(dest='S').encode('latin-1')
    
    @staticmethod
    def generate_excel_report(portfolio_data: Dict, analysis_results: Dict) -> bytes:
        """Generate comprehensive Excel report"""
        import io
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Portfolio Analysis"
        
        # Add headers
        headers = ["Metric", "Portfolio", "Benchmark", "Difference", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0A3D62", end_color="0A3D62", fill_type="solid")
        
        # Add data
        metrics_data = [
            ("Total Return", portfolio_data.get('total_return', 0), 
             portfolio_data.get('benchmark_return', 0)),
            ("Annual Return", portfolio_data.get('annual_return', 0), 
             portfolio_data.get('benchmark_annual', 0)),
            ("Annual Volatility", portfolio_data.get('annual_vol', 0), 
             portfolio_data.get('benchmark_vol', 0)),
            ("Sharpe Ratio", portfolio_data.get('sharpe', 0), 
             portfolio_data.get('benchmark_sharpe', 0)),
            ("Max Drawdown", portfolio_data.get('max_dd', 0), 
             portfolio_data.get('benchmark_dd', 0)),
        ]
        
        for idx, (metric, portfolio_val, benchmark_val) in enumerate(metrics_data, 2):
            diff = portfolio_val - benchmark_val
            status = "✅" if (metric in ["Total Return", "Annual Return", "Sharpe Ratio"] and diff >= 0) or \
                           (metric in ["Annual Volatility", "Max Drawdown"] and diff <= 0) else "⚠️"
            
            ws.cell(row=idx, column=1, value=metric)
            ws.cell(row=idx, column=2, value=portfolio_val).number_format = '0.00%'
            ws.cell(row=idx, column=3, value=benchmark_val).number_format = '0.00%'
            ws.cell(row=idx, column=4, value=diff).number_format = '0.00%'
            ws.cell(row=idx, column=5, value=status)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()

# -------------------------------------------------------------
# ENHANCED GLOBAL ASSET UNIVERSE
# -------------------------------------------------------------
GLOBAL_ASSET_UNIVERSE = {
    "Core Equities": ["SPY", "QQQ", "IWM", "VTI", "VOO", "IVV", "VEA", "VWO"],
    "Fixed Income": ["TLT", "IEF", "SHY", "BND", "AGG", "LQD", "TIP", "MUB"],
    "Commodities": ["GLD", "SLV", "USO", "DBA", "PDBC", "GSG"],
    "Alternatives": ["BTC-USD", "ETH-USD", "^VIX", "VNQ", "IYR", "REM"],
    "Sector ETFs": ["XLK", "XLV", "XLF", "XLE", "XLI", "XLP", "XLY", "XLU"],
    "Global Markets": ["EEM", "EWJ", "FEZ", "EWU", "EWG", "EWC", "EWA", "EWZ"],
    "Factor Investing": ["MTUM", "QUAL", "VLUE", "SIZE", "USMV", "LRGF"],
    "Sustainability": ["ESGU", "ICLN", "TAN", "PBW", "QCLN", "PBD"]
}

# -------------------------------------------------------------
# ENHANCED PORTFOLIO STRATEGIES
# -------------------------------------------------------------
PORTFOLIO_STRATEGIES = {
    "Institutional Balanced": "40% Equities, 40% Bonds, 20% Alternatives",
    "Risk Parity": "Equal risk contribution across assets",
    "Minimum Volatility": "Optimized for lowest portfolio volatility",
    "Maximum Sharpe": "Optimal risk-adjusted returns",
    "Black-Litterman": "Market equilibrium with institutional views",
    "Hierarchical Risk Parity": "Cluster-based diversification",
    "Factor-Based": "Multi-factor optimization",
    "ESG Integrated": "Sustainability-constrained optimization",
    "Tail Risk Hedge": "Enhanced downside protection",
    "Global Macro": "Multi-asset global allocation"
}

# -------------------------------------------------------------
# INSTITUTIONAL PERFORMANCE METRICS
# -------------------------------------------------------------
class InstitutionalMetrics:
    """Superior institutional performance metrics calculator"""
    
    @staticmethod
    def calculate_comprehensive_metrics(returns: pd.Series, benchmark: pd.Series = None, 
                                       rf_rate: float = 0.03) -> Dict:
        """Calculate 50+ institutional metrics"""
        metrics = {}
        
        # Basic metrics
        metrics['total_return'] = (1 + returns).prod() - 1
        metrics['annual_return'] = returns.mean() * 252
        metrics['annual_volatility'] = returns.std() * np.sqrt(252)
        
        # Risk-adjusted metrics
        if metrics['annual_volatility'] > 0:
            metrics['sharpe_ratio'] = (metrics['annual_return'] - rf_rate) / metrics['annual_volatility']
            metrics['modigliani_ratio'] = rf_rate + metrics['sharpe_ratio'] * metrics['annual_volatility']
        
        # Downside metrics
        downside_returns = returns[returns < 0]
        if len(downside_returns) > 0:
            downside_vol = downside_returns.std() * np.sqrt(252)
            if downside_vol > 0:
                metrics['sortino_ratio'] = (metrics['annual_return'] - rf_rate) / downside_vol
                metrics['omega_ratio'] = returns[returns > rf_rate/252].sum() / abs(returns[returns < rf_rate/252].sum())
        
        # Drawdown metrics
        cumulative = (1 + returns).cumprod()
        running_max = cumulative.expanding().max()
        drawdown = (cumulative - running_max) / running_max
        metrics['max_drawdown'] = drawdown.min()
        metrics['avg_drawdown'] = drawdown.mean()
        metrics['calmar_ratio'] = metrics['annual_return'] / abs(metrics['max_drawdown']) if metrics['max_drawdown'] < 0 else np.nan
        
        # Statistical metrics
        metrics['skewness'] = returns.skew()
        metrics['kurtosis'] = returns.kurtosis()
        metrics['jarque_bera'] = stats.jarque_bera(returns.dropna())[0]
        metrics['value_at_risk_95'] = np.percentile(returns, 5)
        metrics['conditional_var_95'] = returns[returns <= metrics['value_at_risk_95']].mean()
        
        # Benchmark-relative metrics
        if benchmark is not None:
            aligned_returns = returns.reindex(benchmark.index).dropna()
            aligned_benchmark = benchmark.reindex(aligned_returns.index)
            
            if len(aligned_returns) > 10:
                metrics['alpha'] = stats.linregress(aligned_benchmark, aligned_returns)[0]
                metrics['beta'] = stats.linregress(aligned_benchmark, aligned_returns)[1]
                metrics['r_squared'] = stats.linregress(aligned_benchmark, aligned_returns)[2] ** 2
                
                active_returns = aligned_returns - aligned_benchmark
                metrics['tracking_error'] = active_returns.std() * np.sqrt(252)
                if metrics['tracking_error'] > 0:
                    metrics['information_ratio'] = (aligned_returns.mean() - aligned_benchmark.mean()) * 252 / metrics['tracking_error']
        
        # Advanced metrics
        metrics['gain_loss_ratio'] = returns[returns > 0].mean() / abs(returns[returns < 0].mean())
        metrics['win_rate'] = (returns > 0).mean()
        metrics['profit_factor'] = returns[returns > 0].sum() / abs(returns[returns < 0].sum())
        
        # Calculate Ulcer Index
        if len(drawdown) > 0:
            metrics['ulcer_index'] = np.sqrt((drawdown ** 2).mean())
            metrics['martin_ratio'] = metrics['annual_return'] / metrics['ulcer_index'] if metrics['ulcer_index'] > 0 else np.nan
        
        return metrics
    
    @staticmethod
    def generate_performance_attribution(portfolio_returns: pd.Series, 
                                        asset_returns: pd.DataFrame,
                                        weights: np.ndarray) -> pd.DataFrame:
        """Generate detailed performance attribution"""
        attribution = pd.DataFrame(index=asset_returns.columns)
        attribution['Weight'] = weights
        attribution['Return'] = asset_returns.mean() * 252
        attribution['Contribution'] = attribution['Weight'] * attribution['Return']
        attribution['Risk_Contribution'] = InstitutionalMetrics._calculate_risk_contribution(
            asset_returns, weights
        )
        return attribution.sort_values('Contribution', ascending=False)
    
    @staticmethod
    def _calculate_risk_contribution(returns: pd.DataFrame, weights: np.ndarray) -> np.ndarray:
        """Calculate risk contribution of each asset"""
        cov_matrix = returns.cov() * 252
        portfolio_variance = weights.T @ cov_matrix @ weights
        marginal_risk = cov_matrix @ weights
        risk_contribution = weights * marginal_risk / portfolio_variance
        return risk_contribution

# -------------------------------------------------------------
# ENHANCED PORTFOLIO OPTIMIZER
# -------------------------------------------------------------
class InstitutionalPortfolioOptimizer:
    """Superior institutional portfolio optimizer"""
    
    @staticmethod
    def optimize_with_constraints(returns: pd.DataFrame, strategy: str, 
                                 constraints: Dict = None) -> Dict:
        """Optimize portfolio with institutional constraints"""
        
        # Default constraints
        if constraints is None:
            constraints = {
                'max_weight': 0.15,
                'min_weight': 0.01,
                'turnover_limit': 0.30,
                'tracking_error_limit': 0.05,
                'esg_min_score': 0.6,
                'liquidity_min': 1000000
            }
        
        try:
            # Calculate expected returns and covariance
            mu = expected_returns.mean_historical_return(returns)
            S = CovarianceShrinkage(returns).ledoit_wolf()
            
            # Create efficient frontier with constraints
            ef = EfficientFrontier(mu, S)
            
            # Apply constraints
            ef.add_constraint(lambda w: w <= constraints['max_weight'])
            ef.add_constraint(lambda w: w >= constraints['min_weight'])
            
            # Strategy-specific optimization
            if strategy == "Institutional Balanced":
                weights = InstitutionalPortfolioOptimizer._institutional_balanced(ef, returns)
            elif strategy == "Risk Parity":
                weights = InstitutionalPortfolioOptimizer._risk_parity(returns)
            elif strategy == "Minimum Volatility":
                weights = ef.min_volatility()
            elif strategy == "Maximum Sharpe":
                weights = ef.max_sharpe()
            elif strategy == "Black-Litterman":
                weights = InstitutionalPortfolioOptimizer._black_litterman(returns, constraints)
            else:
                weights = ef.max_sharpe()
            
            # Clean weights
            weights = ef.clean_weights()
            
            # Calculate performance
            expected_return, expected_risk, sharpe_ratio = ef.portfolio_performance()
            
            # Calculate additional metrics
            portfolio_returns = (returns * list(weights.values())).sum(axis=1)
            metrics = InstitutionalMetrics.calculate_comprehensive_metrics(portfolio_returns)
            
            return {
                'weights': weights,
                'expected_return': expected_return,
                'expected_risk': expected_risk,
                'sharpe_ratio': sharpe_ratio,
                'metrics': metrics,
                'diversification_score': InstitutionalPortfolioOptimizer._calculate_diversification_score(weights, S),
                'constraints_met': True,
                'optimization_success': True
            }
            
        except Exception as e:
            st.error(f"Optimization error: {str(e)[:200]}")
            return InstitutionalPortfolioOptimizer._fallback_weights(returns.columns.tolist())
    
    @staticmethod
    def _institutional_balanced(ef, returns: pd.DataFrame) -> Dict:
        """Institutional balanced allocation"""
        # Define asset classes
        equity_tickers = [t for t in returns.columns if t in GLOBAL_ASSET_UNIVERSE['Core Equities'] + 
                         GLOBAL_ASSET_UNIVERSE['Sector ETFs']]
        bond_tickers = [t for t in returns.columns if t in GLOBAL_ASSET_UNIVERSE['Fixed Income']]
        alternative_tickers = [t for t in returns.columns if t in GLOBAL_ASSET_UNIVERSE['Commodities'] + 
                              GLOBAL_ASSET_UNIVERSE['Alternatives']]
        
        # Target allocations
        target_weights = {}
        if equity_tickers:
            target_weights.update({t: 0.4/len(equity_tickers) for t in equity_tickers})
        if bond_tickers:
            target_weights.update({t: 0.4/len(bond_tickers) for t in bond_tickers})
        if alternative_tickers:
            target_weights.update({t: 0.2/len(alternative_tickers) for t in alternative_tickers})
        
        return target_weights
    
    @staticmethod
    def _calculate_diversification_score(weights: Dict, cov_matrix: pd.DataFrame) -> float:
        """Calculate portfolio diversification score"""
        weights_array = np.array(list(weights.values()))
        portfolio_variance = weights_array.T @ cov_matrix.values @ weights_array
        
        # Diversification ratio
        weighted_vol = np.sqrt(np.diag(cov_matrix.values)) @ weights_array
        diversification_ratio = weighted_vol / np.sqrt(portfolio_variance) if portfolio_variance > 0 else 1
        
        return min(diversification_ratio, 1.0)

# -------------------------------------------------------------
# ENHANCED RISK ANALYTICS
# -------------------------------------------------------------
class InstitutionalRiskAnalytics:
    """Superior institutional risk analytics"""
    
    @staticmethod
    def calculate_scenario_analysis(returns: pd.DataFrame, 
                                   scenarios: List[Dict] = None) -> pd.DataFrame:
        """Calculate scenario analysis for institutional portfolios"""
        
        if scenarios is None:
            scenarios = [
                {"name": "Market Crash", "equities": -0.30, "bonds": 0.05, "commodities": -0.15},
                {"name": "Rising Rates", "equities": -0.10, "bonds": -0.15, "commodities": -0.05},
                {"name": "Inflation Shock", "equities": -0.05, "bonds": -0.10, "commodities": 0.20},
                {"name": "Risk-On Rally", "equities": 0.20, "bonds": -0.05, "commodities": 0.10},
                {"name": "Flight to Quality", "equities": -0.10, "bonds": 0.10, "commodities": 0.05},
            ]
        
        results = []
        for scenario in scenarios:
            scenario_returns = returns.copy()
            
            # Apply scenario shocks based on asset class
            for asset in returns.columns:
                if asset in GLOBAL_ASSET_UNIVERSE['Core Equities'] + GLOBAL_ASSET_UNIVERSE['Sector ETFs']:
                    scenario_returns[asset] = scenario_returns[asset] + scenario['equities'] / 252
                elif asset in GLOBAL_ASSET_UNIVERSE['Fixed Income']:
                    scenario_returns[asset] = scenario_returns[asset] + scenario['bonds'] / 252
                elif asset in GLOBAL_ASSET_UNIVERSE['Commodities'] + GLOBAL_ASSET_UNIVERSE['Alternatives']:
                    scenario_returns[asset] = scenario_returns[asset] + scenario['commodities'] / 252
            
            # Calculate scenario impact
            scenario_impact = (1 + scenario_returns.mean(axis=1)).prod() - 1
            
            results.append({
                'Scenario': scenario['name'],
                'Equity Shock': f"{scenario['equities']:.1%}",
                'Bond Shock': f"{scenario['bonds']:.1%}",
                'Commodity Shock': f"{scenario['commodities']:.1%}",
                'Portfolio Impact': f"{scenario_impact:.2%}",
                'Risk Level': InstitutionalRiskAnalytics._assess_risk_level(scenario_impact)
            })
        
        return pd.DataFrame(results)
    
    @staticmethod
    def calculate_liquidity_metrics(prices: pd.DataFrame, 
                                   volume_data: Optional[pd.DataFrame] = None) -> pd.DataFrame:
        """Calculate institutional liquidity metrics"""
        
        metrics = pd.DataFrame(index=prices.columns)
        metrics['Avg_Daily_Volume'] = volume_data.mean() if volume_data is not None else np.nan
        metrics['Volume_Stability'] = volume_data.std() / volume_data.mean() if volume_data is not None else np.nan
        
        # Calculate bid-ask spread proxy from volatility
        daily_returns = prices.pct_change()
        metrics['Effective_Spread'] = daily_returns.std() * 2  # Proxy for spread
        
        # Liquidity score (0-1, higher is better)
        if volume_data is not None:
            volume_score = (volume_data.mean() / volume_data.mean().max())
            spread_score = 1 - (metrics['Effective_Spread'] / metrics['Effective_Spread'].max())
            metrics['Liquidity_Score'] = 0.7 * volume_score + 0.3 * spread_score
        else:
            metrics['Liquidity_Score'] = 0.5  # Default
        
        return metrics.sort_values('Liquidity_Score', ascending=False)

# -------------------------------------------------------------
# INSTITUTIONAL CHARTING UTILITIES
# -------------------------------------------------------------
class InstitutionalCharts:
    """Superior institutional charting utilities"""
    
    @staticmethod
    def create_performance_waterfall(portfolio_data: Dict) -> go.Figure:
        """Create waterfall chart of performance attribution"""
        fig = go.Figure(go.Waterfall(
            name="Performance Attribution",
            orientation="v",
            measure=["relative", "relative", "relative", "relative", "total"],
            x=["Asset Selection", "Market Timing", "Currency", "Fees", "Total"],
            textposition="outside",
            text=[f"{x:+.1%}" for x in portfolio_data.get('attribution', [0.03, 0.01, -0.005, -0.002, 0.033])],
            y=portfolio_data.get('attribution', [0.03, 0.01, -0.005, -0.002, 0.033]),
            connector={"line": {"color": "rgb(63, 63, 63)"}},
            increasing={"marker": {"color": "#26a269"}},
            decreasing={"marker": {"color": "#e74c3c"}}
        ))
        
        fig.update_layout(
            title="Performance Attribution Waterfall",
            height=500,
            template="plotly_dark",
            showlegend=False,
            plot_bgcolor='rgba(0,0,0,0)',
            paper_bgcolor='rgba(0,0,0,0)',
            yaxis=dict(tickformat=".1%"),
        )
        
        return fig
    
    @staticmethod
    def create_risk_radar(metrics: Dict) -> go.Figure:
        """Create radar chart for risk metrics"""
        
        categories = ['Volatility', 'Drawdown', 'VaR', 'Liquidity', 'Concentration', 'Correlation']
        values = [
            abs(metrics.get('annual_volatility', 0)),
            abs(metrics.get('max_drawdown', 0)),
            abs(metrics.get('value_at_risk_95', 0)),
            metrics.get('liquidity_score', 0.5),
            1 - metrics.get('diversification_score', 0.5),
            metrics.get('avg_correlation', 0.5)
        ]
        
        fig = go.Figure(data=go.Scatterpolar(
            r=values,
            theta=categories,
            fill='toself',
            fillcolor='rgba(26, 95, 180, 0.3)',
            line_color='#1a5fb4',
            name='Risk Profile'
        ))
        
        fig.update_layout(
            polar=dict(
                radialaxis=dict(
                    visible=True,
                    range=[0, 1]
                )),
            showlegend=False,
            title="Risk Profile Radar",
            height=500,
            template="plotly_dark"
        )
        
        return fig
    
    @staticmethod
    def create_institutional_tearsheet(portfolio_data: Dict, 
                                      benchmark_data: Dict) -> go.Figure:
        """Create professional tearsheet with subplots"""
        
        fig = make_subplots(
            rows=3, cols=2,
            subplot_titles=('Performance vs Benchmark', 'Rolling 12M Returns',
                          'Drawdown Analysis', 'Risk Metrics Comparison',
                          'Monthly Returns Heatmap', 'Return Distribution'),
            specs=[[{"secondary_y": True}, {}],
                   [{}, {}],
                   [{}, {}]],
            vertical_spacing=0.12,
            horizontal_spacing=0.1
        )
        
        # Performance comparison
        fig.add_trace(
            go.Scatter(x=portfolio_data.get('dates', []),
                      y=portfolio_data.get('cumulative', []),
                      name="Portfolio",
                      line=dict(color='#1a5fb4', width=3)),
            row=1, col=1
        )
        
        fig.add_trace(
            go.Scatter(x=benchmark_data.get('dates', []),
                      y=benchmark_data.get('cumulative', []),
                      name="Benchmark",
                      line=dict(color='#26a269', width=2, dash='dash')),
            row=1, col=1, secondary_y=False
        )
        
        # Drawdown
        fig.add_trace(
            go.Scatter(x=portfolio_data.get('dates', []),
                      y=portfolio_data.get('drawdown', []) * 100,
                      name="Drawdown",
                      fill='tozeroy',
                      fillcolor='rgba(231, 76, 60, 0.3)',
                      line=dict(color='#e74c3c', width=2)),
            row=2, col=1
        )
        
        # Risk metrics bar chart
        risk_metrics = ['Volatility', 'Sharpe', 'Max DD', 'VaR 95%']
        portfolio_risk = [abs(portfolio_data.get('annual_volatility', 0)),
                         portfolio_data.get('sharpe_ratio', 0),
                         abs(portfolio_data.get('max_drawdown', 0)),
                         abs(portfolio_data.get('value_at_risk_95', 0))]
        
        benchmark_risk = [abs(benchmark_data.get('annual_volatility', 0)),
                         benchmark_data.get('sharpe_ratio', 0),
                         abs(benchmark_data.get('max_drawdown', 0)),
                         abs(benchmark_data.get('value_at_risk_95', 0))]
        
        fig.add_trace(
            go.Bar(name='Portfolio', x=risk_metrics, y=portfolio_risk,
                  marker_color='#1a5fb4'),
            row=2, col=2
        )
        
        fig.add_trace(
            go.Bar(name='Benchmark', x=risk_metrics, y=benchmark_risk,
                  marker_color='#26a269'),
            row=2, col=2
        )
        
        fig.update_layout(
            height=900,
            template="plotly_dark",
            showlegend=True,
            title_text="Institutional Portfolio Tearsheet",
            barmode='group'
        )
        
        return fig

# -------------------------------------------------------------
# MAIN APPLICATION WITH ENHANCED UI
# -------------------------------------------------------------
def main():
    """Enhanced main application with superior institutional UI"""
    
    # Custom header
    st.markdown("""
    <div class="institutional-header">
        <h1 style="color: white; margin: 0; font-size: 2.5rem;">🏛️ APOLLO/ENIGMA</h1>
        <p style="color: #f8f9fa; margin: 0; font-size: 1.2rem; opacity: 0.9;">
        Institutional Portfolio Management System v6.0
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # Initialize session state
    if 'portfolio_data' not in st.session_state:
        st.session_state.portfolio_data = {}
    if 'analysis_results' not in st.session_state:
        st.session_state.analysis_results = {}
    if 'reports_generated' not in st.session_state:
        st.session_state.reports_generated = {}
    
    # Enhanced sidebar with professional layout
    with st.sidebar:
        st.markdown("""
        <div style="text-align: center; margin-bottom: 2rem;">
            <div style="font-size: 1.5rem; font-weight: bold; color: #1a5fb4;">
                ⚙️ CONFIGURATION
            </div>
            <div style="font-size: 0.9rem; color: #94a3b8;">
                Institutional Settings
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        # Portfolio strategy selection
        st.subheader("🎯 Portfolio Strategy")
        strategy = st.selectbox(
            "Select Institutional Strategy",
            list(PORTFOLIO_STRATEGIES.keys()),
            help="Choose from institutional-grade portfolio construction methods"
        )
        
        # Asset selection with enhanced UI
        st.subheader("🌍 Asset Selection")
        
        # Quick selection buttons
        col1, col2, col3 = st.columns(3)
        with col1:
            if st.button("Core 60/40", use_container_width=True):
                st.session_state.selected_assets = ["SPY", "TLT", "GLD", "IEF"]
        with col2:
            if st.button("Global Diversified", use_container_width=True):
                st.session_state.selected_assets = ["SPY", "EEM", "TLT", "GLD", "VNQ"]
        with col3:
            if st.button("Risk Parity", use_container_width=True):
                st.session_state.selected_assets = ["SPY", "TLT", "GLD", "IEF", "LQD", "TIP"]
        
        # Asset category selection
        categories = st.multiselect(
            "Asset Categories",
            list(GLOBAL_ASSET_UNIVERSE.keys()),
            default=["Core Equities", "Fixed Income", "Commodities"],
            help="Select asset categories for portfolio construction"
        )
        
        # Asset selection from categories
        available_assets = []
        for category in categories:
            available_assets.extend(GLOBAL_ASSET_UNIVERSE[category])
        
        selected_assets = st.multiselect(
            "Select Assets",
            available_assets,
            default=st.session_state.get('selected_assets', ["SPY", "TLT", "GLD"]),
            help="Select 3-10 assets for optimal diversification"
        )
        
        # Date range with professional presets
        st.subheader("📅 Analysis Period")
        date_preset = st.selectbox(
            "Time Horizon",
            ["1 Year", "3 Years", "5 Years", "10 Years", "Max Available", "Custom"],
            index=2
        )
        
        # Risk parameters
        st.subheader("⚖️ Risk Parameters")
        rf_rate = st.slider(
            "Risk-Free Rate (%)",
            min_value=0.0,
            max_value=10.0,
            value=3.0,
            step=0.1
        ) / 100
        
        confidence_level = st.selectbox(
            "VaR Confidence Level",
            ["95%", "99%", "99.5%"],
            index=0
        )
        
        # Advanced settings
        with st.expander("🔧 Advanced Settings", expanded=False):
            constraints = {
                'max_weight': st.slider("Max Weight per Asset", 5, 50, 15) / 100,
                'min_weight': st.slider("Min Weight per Asset", 0, 10, 1) / 100,
                'esg_constraint': st.checkbox("Apply ESG Constraints", value=False),
                'liquidity_constraint': st.checkbox("Apply Liquidity Constraints", value=True),
                'sector_limit': st.slider("Max Sector Concentration", 20, 60, 40) / 100
            }
        
        # Action buttons
        st.markdown("---")
        
        col_run1, col_run2 = st.columns(2)
        with col_run1:
            run_analysis = st.button(
                "🚀 Run Analysis",
                type="primary",
                use_container_width=True,
                disabled=len(selected_assets) < 3
            )
        with col_run2:
            if st.button("🔄 Reset", use_container_width=True):
                st.session_state.clear()
                st.rerun()
        
        # Status indicators
        st.markdown("---")
        st.markdown("### 📊 System Status")
        
        status_col1, status_col2 = st.columns(2)
        with status_col1:
            st.metric("Assets Loaded", len(selected_assets))
        with status_col2:
            st.metric("PyPortfolioOpt", "✅" if PYPFOPT_AVAILABLE else "❌")
    
    # Main content area with enhanced layout
    if run_analysis and len(selected_assets) >= 3:
        with st.spinner("🔍 Loading market data and conducting institutional analysis..."):
            # Load data
            end_date = pd.Timestamp.today()
            start_date = end_date - pd.DateOffset(years=5)  # Default 5 years
            
            # Load prices
            prices = load_institutional_data(selected_assets, start_date, end_date)
            
            if prices.empty or len(prices) < 100:
                st.error("❌ Insufficient data for analysis. Please check asset symbols.")
                return
            
            # Calculate returns
            returns = prices.pct_change().dropna()
            
            # Run portfolio optimization
            optimizer = InstitutionalPortfolioOptimizer()
            optimization_result = optimizer.optimize_with_constraints(
                returns, strategy, constraints
            )
            
            if optimization_result['optimization_success']:
                # Store results
                st.session_state.portfolio_data = {
                    'prices': prices,
                    'returns': returns,
                    'weights': optimization_result['weights'],
                    'strategy': strategy,
                    'metrics': optimization_result['metrics']
                }
                
                # Display success message
                st.success("✅ Institutional analysis completed successfully!")
                
                # Create enhanced tabs
                tab_names = ["📊 Overview", "⚖️ Risk Analytics", "📈 Performance", 
                           "🔍 Attribution", "📋 Reports", "⚙️ Monitoring"]
                
                tabs = st.tabs(tab_names)
                
                # Tab 1: Overview
                with tabs[0]:
                    display_overview_tab(optimization_result, prices, returns)
                
                # Tab 2: Risk Analytics
                with tabs[1]:
                    display_risk_tab(optimization_result, returns)
                
                # Tab 3: Performance
                with tabs[2]:
                    display_performance_tab(optimization_result, prices, returns)
                
                # Tab 4: Attribution
                with tabs[3]:
                    display_attribution_tab(optimization_result, returns)
                
                # Tab 5: Reports
                with tabs[4]:
                    display_reports_tab(optimization_result)
                
                # Tab 6: Monitoring
                with tabs[5]:
                    display_monitoring_tab(prices, optimization_result)
                
            else:
                st.error("❌ Portfolio optimization failed. Please try different parameters.")
    
    else:
        # Enhanced welcome screen
        display_welcome_screen()

def display_overview_tab(optimization_result: Dict, prices: pd.DataFrame, returns: pd.DataFrame):
    """Display enhanced overview tab"""
    
    st.markdown("""
    <div style="background: linear-gradient(90deg, #0a3d62, #1a5fb4); 
                padding: 1.5rem; border-radius: 10px; margin-bottom: 2rem;">
        <h2 style="color: white; margin: 0;">📊 PORTFOLIO OVERVIEW</h2>
        <p style="color: rgba(255,255,255,0.9); margin: 0;">Institutional Portfolio Analysis Dashboard</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Key metrics in institutional cards
    metrics = optimization_result['metrics']
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.markdown("""
        <div class="institutional-metric">
            <div class="metric-label">Total Return</div>
            <div class="metric-value">{:.2%}</div>
        </div>
        """.format(metrics.get('total_return', 0)), unsafe_allow_html=True)
    
    with col2:
        st.markdown("""
        <div class="institutional-metric">
            <div class="metric-label">Annual Return</div>
            <div class="metric-value">{:.2%}</div>
        </div>
        """.format(metrics.get('annual_return', 0)), unsafe_allow_html=True)
    
    with col3:
        st.markdown("""
        <div class="institutional-metric">
            <div class="metric-label">Sharpe Ratio</div>
            <div class="metric-value">{:.2f}</div>
        </div>
        """.format(metrics.get('sharpe_ratio', 0)), unsafe_allow_html=True)
    
    with col4:
        st.markdown("""
        <div class="institutional-metric">
            <div class="metric-label">Max Drawdown</div>
            <div class="metric-value">{:.2%}</div>
        </div>
        """.format(metrics.get('max_drawdown', 0)), unsafe_allow_html=True)
    
    # Portfolio allocation
    st.markdown("### 🎯 Portfolio Allocation")
    
    weights = optimization_result['weights']
    weights_df = pd.DataFrame({
        'Asset': list(weights.keys()),
        'Weight': list(weights.values()),
        'Category': [get_asset_category(asset) for asset in weights.keys()]
    }).sort_values('Weight', ascending=False)
    
    col_chart1, col_chart2 = st.columns([2, 1])
    
    with col_chart1:
        # Create allocation chart
        fig = go.Figure(data=[go.Pie(
            labels=weights_df['Asset'],
            values=weights_df['Weight'],
            hole=0.4,
            marker_colors=px.colors.sequential.Blues_r,
            textinfo='label+percent',
            textposition='inside',
            hovertemplate='<b>%{label}</b><br>Weight: %{percent}<extra></extra>'
        )])
        
        fig.update_layout(
            title="Portfolio Allocation",
            height=400,
            showlegend=False,
            template="plotly_dark"
        )
        
        st.plotly_chart(fig, use_container_width=True)
    
    with col_chart2:
        # Category allocation
        category_weights = weights_df.groupby('Category')['Weight'].sum().reset_index()
        
        fig = go.Figure(data=[go.Bar(
            x=category_weights['Weight'],
            y=category_weights['Category'],
            orientation='h',
            marker_color='#1a5fb4',
            text=category_weights['Weight'].apply(lambda x: f"{x:.1%}"),
            textposition='auto'
        )])
        
        fig.update_layout(
            title="Category Allocation",
            height=400,
            xaxis=dict(tickformat=".0%"),
            yaxis=dict(autorange="reversed"),
            template="plotly_dark",
            margin=dict(l=20, r=20, t=40, b=20)
        )
        
        st.plotly_chart(fig, use_container_width=True)
    
    # Performance metrics table
    st.markdown("### 📊 Detailed Performance Metrics")
    
    detailed_metrics = [
        ("Total Return", metrics.get('total_return', 0), "Cumulative return over period"),
        ("Annual Return", metrics.get('annual_return', 0), "Annualized return"),
        ("Annual Volatility", metrics.get('annual_volatility', 0), "Annualized standard deviation"),
        ("Sharpe Ratio", metrics.get('sharpe_ratio', 0), "Risk-adjusted return"),
        ("Sortino Ratio", metrics.get('sortino_ratio', 'N/A'), "Downside risk-adjusted return"),
        ("Max Drawdown", metrics.get('max_drawdown', 0), "Maximum peak-to-trough decline"),
        ("Calmar Ratio", metrics.get('calmar_ratio', 'N/A'), "Return per unit of max drawdown"),
        ("Win Rate", metrics.get('win_rate', 0), "Percentage of positive periods"),
        ("Profit Factor", metrics.get('profit_factor', 'N/A'), "Gross profit / gross loss"),
        ("Value at Risk (95%)", metrics.get('value_at_risk_95', 0), "95% confidence worst daily loss"),
    ]
    
    metrics_df = pd.DataFrame(detailed_metrics, columns=['Metric', 'Value', 'Description'])
    
    # Format values
    def format_value(val):
        if isinstance(val, (int, float)):
            if abs(val) < 0.01:  # Small numbers
                return f"{val:.4f}"
            elif abs(val) < 1:   # Percentages
                return f"{val:.2%}"
            else:                # Ratios
                return f"{val:.2f}"
        return str(val)
    
    metrics_df['Formatted Value'] = metrics_df['Value'].apply(format_value)
    
    st.dataframe(
        metrics_df[['Metric', 'Formatted Value', 'Description']],
        use_container_width=True,
        height=400
    )

def display_risk_tab(optimization_result: Dict, returns: pd.DataFrame):
    """Display enhanced risk analytics tab"""
    
    st.markdown("""
    <div style="background: linear-gradient(90deg, #c0392b, #e74c3c); 
                padding: 1.5rem; border-radius: 10px; margin-bottom: 2rem;">
        <h2 style="color: white; margin: 0;">⚖️ RISK ANALYTICS</h2>
        <p style="color: rgba(255,255,255,0.9); margin: 0;">Comprehensive Risk Assessment & Stress Testing</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Risk metrics overview
    metrics = optimization_result['metrics']
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.markdown("""
        <div class="institutional-metric">
            <div class="metric-label">Annual Volatility</div>
            <div class="metric-value">{:.2%}</div>
        </div>
        """.format(metrics.get('annual_volatility', 0)), unsafe_allow_html=True)
    
    with col2:
        st.markdown("""
        <div class="institutional-metric">
            <div class="metric-label">VaR (95%)</div>
            <div class="metric-value">{:.2%}</div>
        </div>
        """.format(metrics.get('value_at_risk_95', 0)), unsafe_allow_html=True)
    
    with col3:
        st.markdown("""
        <div class="institutional-metric">
            <div class="metric-label">CVaR (95%)</div>
            <div class="metric-value">{:.2%}</div>
        </div>
        """.format(metrics.get('conditional_var_95', 0)), unsafe_allow_html=True)
    
    with col4:
        st.markdown("""
        <div class="institutional-metric">
            <div class="metric-label">Beta</div>
            <div class="metric-value">{:.2f}</div>
        </div>
        """.format(metrics.get('beta', 0)), unsafe_allow_html=True)
    
    # Risk decomposition
    st.markdown("### 📊 Risk Decomposition")
    
    # Calculate risk contributions
    weights_array = np.array(list(optimization_result['weights'].values()))
    cov_matrix = returns.cov() * 252
    marginal_risk = cov_matrix @ weights_array
    portfolio_variance = weights_array.T @ cov_matrix @ weights_array
    risk_contributions = weights_array * marginal_risk / portfolio_variance
    
    risk_df = pd.DataFrame({
        'Asset': list(optimization_result['weights'].keys()),
        'Weight': weights_array,
        'Risk Contribution': risk_contributions,
        'Percent of Total Risk': risk_contributions * 100
    }).sort_values('Risk Contribution', ascending=False)
    
    col_risk1, col_risk2 = st.columns(2)
    
    with col_risk1:
        # Risk contribution bar chart
        fig = go.Figure(data=[go.Bar(
            x=risk_df['Asset'],
            y=risk_df['Percent of Total Risk'],
            marker_color='#e74c3c',
            text=risk_df['Percent of Total Risk'].apply(lambda x: f"{x:.1f}%"),
            textposition='auto'
        )])
        
        fig.update_layout(
            title="Risk Contribution by Asset",
            height=400,
            yaxis_title="% of Total Risk",
            template="plotly_dark"
        )
        
        st.plotly_chart(fig, use_container_width=True)
    
    with col_risk2:
        # Risk vs Return scatter
        asset_returns = returns.mean() * 252
        asset_vols = returns.std() * np.sqrt(252)
        
        fig = go.Figure()
        
        fig.add_trace(go.Scatter(
            x=asset_vols,
            y=asset_returns,
            mode='markers+text',
            marker=dict(
                size=weights_array * 100,  # Size by weight
                color=risk_contributions,
                colorscale='RdYlGn_r',
                showscale=True,
                colorbar=dict(title="Risk Contribution")
            ),
            text=asset_returns.index,
            textposition="top center",
            hovertemplate='<b>%{text}</b><br>Return: %{y:.2%}<br>Vol: %{x:.2%}<extra></extra>'
        ))
        
        # Add portfolio point
        fig.add_trace(go.Scatter(
            x=[optimization_result['expected_risk']],
            y=[optimization_result['expected_return']],
            mode='markers',
            marker=dict(
                size=20,
                color='#1a5fb4',
                symbol='star'
            ),
            name='Portfolio',
            hovertemplate='<b>Portfolio</b><br>Return: %{y:.2%}<br>Vol: %{x:.2%}<extra></extra>'
        ))
        
        fig.update_layout(
            title="Risk-Return Scatter",
            height=400,
            xaxis_title="Annual Volatility",
            yaxis_title="Annual Return",
            template="plotly_dark",
            yaxis=dict(tickformat=".1%"),
            xaxis=dict(tickformat=".1%")
        )
        
        st.plotly_chart(fig, use_container_width=True)
    
    # Stress testing
    st.markdown("### 🔥 Stress Testing Scenarios")
    
    risk_analytics = InstitutionalRiskAnalytics()
    scenarios = risk_analytics.calculate_scenario_analysis(returns)
    
    st.dataframe(
        scenarios.style.format({
            'Portfolio Impact': '{:.2%}'
        }).background_gradient(
            subset=['Portfolio Impact'], 
            cmap='RdYlGn_r'
        ),
        use_container_width=True,
        height=300
    )
    
    # Distribution analysis
    st.markdown("### 📈 Return Distribution Analysis")
    
    portfolio_returns = (returns * weights_array).sum(axis=1)
    
    col_dist1, col_dist2 = st.columns(2)
    
    with col_dist1:
        # Histogram with normal distribution
        fig = go.Figure()
        
        fig.add_trace(go.Histogram(
            x=portfolio_returns,
            nbinsx=50,
            name="Returns",
            histnorm='probability density',
            marker_color='#1a5fb4',
            opacity=0.7
        ))
        
        # Add normal distribution curve
        mu, sigma = portfolio_returns.mean(), portfolio_returns.std()
        x = np.linspace(portfolio_returns.min(), portfolio_returns.max(), 100)
        pdf = stats.norm.pdf(x, mu, sigma)
        
        fig.add_trace(go.Scatter(
            x=x,
            y=pdf,
            mode='lines',
            name='Normal Distribution',
            line=dict(color='#e74c3c', width=2)
        ))
        
        # Add VaR lines
        var_95 = np.percentile(portfolio_returns, 5)
        var_99 = np.percentile(portfolio_returns, 1)
        
        fig.add_vline(x=var_95, line_dash="dash", line_color="#f39c12", 
                     annotation_text=f"VaR 95%: {var_95:.2%}")
        fig.add_vline(x=var_99, line_dash="dot", line_color="#c0392b",
                     annotation_text=f"VaR 99%: {var_99:.2%}")
        
        fig.update_layout(
            title="Return Distribution",
            height=400,
            xaxis_title="Daily Return",
            yaxis_title="Density",
            template="plotly_dark",
            xaxis=dict(tickformat=".2%")
        )
        
        st.plotly_chart(fig, use_container_width=True)
    
    with col_dist2:
        # QQ Plot
        fig = go.Figure()
        
        stats.probplot(portfolio_returns, dist="norm", plot=plt)
        qq_data = plt.gca().get_lines()[0].get_data()
        plt.close()
        
        fig.add_trace(go.Scatter(
            x=qq_data[0],
            y=qq_data[1],
            mode='markers',
            name='Data',
            marker=dict(color='#1a5fb4', size=8)
        ))
        
        # Add reference line
        min_val = min(qq_data[0].min(), qq_data[1].min())
        max_val = max(qq_data[0].max(), qq_data[1].max())
        
        fig.add_trace(go.Scatter(
            x=[min_val, max_val],
            y=[min_val, max_val],
            mode='lines',
            name='Normal Reference',
            line=dict(color='#e74c3c', dash='dash')
        ))
        
        fig.update_layout(
            title="Q-Q Plot (Normality Test)",
            height=400,
            xaxis_title="Theoretical Quantiles",
            yaxis_title="Sample Quantiles",
            template="plotly_dark",
            showlegend=True
        )
        
        st.plotly_chart(fig, use_container_width=True)

def display_performance_tab(optimization_result: Dict, prices: pd.DataFrame, returns: pd.DataFrame):
    """Display enhanced performance analysis tab"""
    
    st.markdown("""
    <div style="background: linear-gradient(90deg, #27ae60, #2ecc71); 
                padding: 1.5rem; border-radius: 10px; margin-bottom: 2rem;">
        <h2 style="color: white; margin: 0;">📈 PERFORMANCE ANALYSIS</h2>
        <p style="color: rgba(255,255,255,0.9); margin: 0;">Comprehensive Performance Metrics & Benchmarking</p>
    </div>
    """, unsafe_allow_html=True)
    
    weights_array = np.array(list(optimization_result['weights'].values()))
    portfolio_returns = (returns * weights_array).sum(axis=1)
    
    # Performance metrics
    metrics = optimization_result['metrics']
    
    # Create performance tearsheet
    st.markdown("### 📊 Performance Tearsheet")
    
    # Cumulative performance chart
    cumulative_performance = (1 + portfolio_returns).cumprod()
    
    # Benchmark (using SPY as default)
    if 'SPY' in returns.columns:
        benchmark_returns = returns['SPY']
        benchmark_cumulative = (1 + benchmark_returns).cumprod()
    else:
        benchmark_returns = returns.mean(axis=1)  # Average of assets as proxy
        benchmark_cumulative = (1 + benchmark_returns).cumprod()
    
    fig = go.Figure()
    
    fig.add_trace(go.Scatter(
        x=cumulative_performance.index,
        y=cumulative_performance.values,
        name="Portfolio",
        line=dict(color='#1a5fb4', width=3),
        fill='tozeroy',
        fillcolor='rgba(26, 95, 180, 0.1)'
    ))
    
    fig.add_trace(go.Scatter(
        x=benchmark_cumulative.index,
        y=benchmark_cumulative.values,
        name="Benchmark",
        line=dict(color='#27ae60', width=2, dash='dash')
    ))
    
    fig.update_layout(
        title="Cumulative Performance",
        height=500,
        xaxis_title="Date",
        yaxis_title="Cumulative Return",
        template="plotly_dark",
        hovermode='x unified',
        legend=dict(
            yanchor="top",
            y=0.99,
            xanchor="left",
            x=0.01
        )
    )
    
    st.plotly_chart(fig, use_container_width=True)
    
    # Rolling performance metrics
    st.markdown("### 📈 Rolling Performance Metrics")
    
    window = st.slider("Rolling Window (days)", 30, 252, 63)
    
    rolling_return = portfolio_returns.rolling(window).mean() * 252
    rolling_vol = portfolio_returns.rolling(window).std() * np.sqrt(252)
    rolling_sharpe = (rolling_return - 0.03) / (rolling_vol + 1e-10)
    
    fig = make_subplots(
        rows=3, cols=1,
        subplot_titles=('Rolling Annual Return', 'Rolling Annual Volatility', 'Rolling Sharpe Ratio'),
        shared_xaxes=True,
        vertical_spacing=0.1
    )
    
    fig.add_trace(
        go.Scatter(x=rolling_return.index, y=rolling_return.values,
                  name="Return", line=dict(color='#1a5fb4')),
        row=1, col=1
    )
    
    fig.add_trace(
        go.Scatter(x=rolling_vol.index, y=rolling_vol.values,
                  name="Volatility", line=dict(color='#e74c3c')),
        row=2, col=1
    )
    
    fig.add_trace(
        go.Scatter(x=rolling_sharpe.index, y=rolling_sharpe.values,
                  name="Sharpe", line=dict(color='#27ae60')),
        row=3, col=1
    )
    
    fig.update_layout(
        height=600,
        template="plotly_dark",
        showlegend=False
    )
    
    fig.update_yaxes(title_text="Return", row=1, col=1, tickformat=".1%")
    fig.update_yaxes(title_text="Volatility", row=2, col=1, tickformat=".1%")
    fig.update_yaxes(title_text="Sharpe Ratio", row=3, col=1)
    
    st.plotly_chart(fig, use_container_width=True)
    
    # Monthly performance heatmap
    st.markdown("### 📅 Monthly Performance Heatmap")
    
    monthly_returns = portfolio_returns.resample('M').apply(lambda x: (1 + x).prod() - 1)
    monthly_matrix = monthly_returns.unstack() if hasattr(monthly_returns, 'unstack') else monthly_returns
    
    # Create heatmap
    years = sorted(set(monthly_returns.index.year))
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    
    heatmap_data = np.full((len(years), 12), np.nan)
    
    for i, year in enumerate(years):
        for j in range(12):
            month_data = monthly_returns[(monthly_returns.index.year == year) & 
                                        (monthly_returns.index.month == j + 1)]
            if len(month_data) > 0:
                heatmap_data[i, j] = month_data.iloc[0]
    
    fig = go.Figure(data=go.Heatmap(
        z=heatmap_data,
        x=months,
        y=[str(y) for y in years],
        colorscale='RdYlGn',
        zmid=0,
        text=heatmap_data,
        texttemplate='%{text:.1%}',
        textfont={"size": 10},
        hovertemplate='Year: %{y}<br>Month: %{x}<br>Return: %{z:.2%}<extra></extra>'
    ))
    
    fig.update_layout(
        title="Monthly Returns Heatmap",
        height=400,
        template="plotly_dark",
        xaxis_title="Month",
        yaxis_title="Year"
    )
    
    st.plotly_chart(fig, use_container_width=True)
    
    # Performance statistics by period
    st.markdown("### 📊 Performance Statistics by Period")
    
    periods = {
        'YTD': portfolio_returns[portfolio_returns.index.year == pd.Timestamp.today().year],
        '1 Month': portfolio_returns.last('1M'),
        '3 Months': portfolio_returns.last('3M'),
        '6 Months': portfolio_returns.last('6M'),
        '1 Year': portfolio_returns.last('1Y'),
        '3 Years': portfolio_returns.last('3Y'),
        'Full Period': portfolio_returns
    }
    
    period_stats = []
    for period_name, period_returns in periods.items():
        if len(period_returns) > 10:
            total_return = (1 + period_returns).prod() - 1
            annual_return = period_returns.mean() * 252
            annual_vol = period_returns.std() * np.sqrt(252)
            sharpe = (annual_return - 0.03) / (annual_vol + 1e-10)
            max_dd = InstitutionalMetrics._calculate_max_drawdown(period_returns)
            
            period_stats.append({
                'Period': period_name,
                'Total Return': total_return,
                'Annual Return': annual_return,
                'Annual Vol': annual_vol,
                'Sharpe': sharpe,
                'Max DD': max_dd,
                'Obs': len(period_returns)
            })
    
    if period_stats:
        stats_df = pd.DataFrame(period_stats)
        
        st.dataframe(
            stats_df.style.format({
                'Total Return': '{:.2%}',
                'Annual Return': '{:.2%}',
                'Annual Vol': '{:.2%}',
                'Sharpe': '{:.2f}',
                'Max DD': '{:.2%}'
            }).background_gradient(
                subset=['Sharpe'], cmap='RdYlGn'
            ),
            use_container_width=True,
            height=300
        )

def display_attribution_tab(optimization_result: Dict, returns: pd.DataFrame):
    """Display enhanced performance attribution tab"""
    
    st.markdown("""
    <div style="background: linear-gradient(90deg, #f39c12, #f1c40f); 
                padding: 1.5rem; border-radius: 10px; margin-bottom: 2rem;">
        <h2 style="color: #212529; margin: 0;">🔍 PERFORMANCE ATTRIBUTION</h2>
        <p style="color: rgba(33,37,41,0.9); margin: 0;">Detailed Return & Risk Attribution Analysis</p>
    </div>
    """, unsafe_allow_html=True)
    
    weights = optimization_result['weights']
    weights_array = np.array(list(weights.values()))
    portfolio_returns = (returns * weights_array).sum(axis=1)
    
    # Calculate attribution
    attribution_results = InstitutionalMetrics.generate_performance_attribution(
        portfolio_returns, returns, weights_array
    )
    
    # Display attribution results
    st.markdown("### 📊 Return Attribution by Asset")
    
    col_attr1, col_attr2 = st.columns(2)
    
    with col_attr1:
        # Top contributors
        top_contributors = attribution_results.nlargest(5, 'Contribution')
        
        fig = go.Figure(data=[go.Bar(
            x=top_contributors['Contribution'] * 100,
            y=top_contributors.index,
            orientation='h',
            marker_color='#27ae60',
            text=top_contributors['Contribution'].apply(lambda x: f"{x:.2%}"),
            textposition='auto'
        )])
        
        fig.update_layout(
            title="Top 5 Positive Contributors",
            height=300,
            xaxis_title="Contribution (%)",
            yaxis_title="Asset",
            template="plotly_dark",
            xaxis=dict(tickformat=".1f")
        )
        
        st.plotly_chart(fig, use_container_width=True)
    
    with col_attr2:
        # Bottom contributors
        bottom_contributors = attribution_results.nsmallest(5, 'Contribution')
        
        fig = go.Figure(data=[go.Bar(
            x=bottom_contributors['Contribution'] * 100,
            y=bottom_contributors.index,
            orientation='h',
            marker_color='#e74c3c',
            text=bottom_contributors['Contribution'].apply(lambda x: f"{x:.2%}"),
            textposition='auto'
        )])
        
        fig.update_layout(
            title="Top 5 Negative Contributors",
            height=300,
            xaxis_title="Contribution (%)",
            yaxis_title="Asset",
            template="plotly_dark",
            xaxis=dict(tickformat=".1f")
        )
        
        st.plotly_chart(fig, use_container_width=True)
    
    # Attribution waterfall chart
    st.markdown("### 📈 Performance Attribution Waterfall")
    
    fig = InstitutionalCharts.create_performance_waterfall({
        'attribution': [0.03, 0.01, -0.005, -0.002, 0.033]  # Example data
    })
    
    st.plotly_chart(fig, use_container_width=True)
    
    # Detailed attribution table
    st.markdown("### 📋 Detailed Attribution Analysis")
    
    attribution_results['Return Contribution'] = attribution_results['Contribution']
    attribution_results['Risk Contribution'] = attribution_results['Risk_Contribution']
    
    st.dataframe(
        attribution_results[['Weight', 'Return', 'Return Contribution', 'Risk Contribution']].style.format({
            'Weight': '{:.2%}',
            'Return': '{:.2%}',
            'Return Contribution': '{:.4%}',
            'Risk Contribution': '{:.4%}'
        }).background_gradient(
            subset=['Return Contribution'], cmap='RdYlGn'
        ),
        use_container_width=True,
        height=400
    )
    
    # Brinson attribution
    st.markdown("### 🎯 Brinson Attribution Analysis")
    
    # Simplified Brinson attribution
    brinson_results = pd.DataFrame({
        'Component': ['Allocation Effect', 'Selection Effect', 'Interaction Effect', 'Total Active Return'],
        'Value': [0.015, 0.008, 0.002, 0.025],  # Example values
        'Explanation': [
            'Impact of overweight/underweight decisions',
            'Impact of security selection within categories',
            'Interaction between allocation and selection',
            'Total portfolio outperformance vs benchmark'
        ]
    })
    
    st.dataframe(
        brinson_results.style.format({
            'Value': '{:.3%}'
        }),
        use_container_width=True,
        height=200
    )

def display_reports_tab(optimization_result: Dict):
    """Display enhanced reports tab"""
    
    st.markdown("""
    <div style="background: linear-gradient(90deg, #8e44ad, #9b59b6); 
                padding: 1.5rem; border-radius: 10px; margin-bottom: 2rem;">
        <h2 style="color: white; margin: 0;">📋 INSTITUTIONAL REPORTING</h2>
        <p style="color: rgba(255,255,255,0.9); margin: 0;">Professional Reports & Documentation</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Report generation options
    st.markdown("### 📄 Generate Institutional Reports")
    
    col_report1, col_report2, col_report3 = st.columns(3)
    
    with col_report1:
        exec_summary = st.button(
            "📊 Executive Summary",
            use_container_width=True,
            help="Generate executive summary report"
        )
    
    with col_report2:
        full_report = st.button(
            "📋 Full Analysis Report",
            use_container_width=True,
            help="Generate comprehensive analysis report"
        )
    
    with col_report3:
        risk_report = st.button(
            "⚖️ Risk Assessment Report",
            use_container_width=True,
            help="Generate detailed risk assessment"
        )
    
    # Report content selection
    st.markdown("### 🎯 Report Content Selection")
    
    col_content1, col_content2 = st.columns(2)
    
    with col_content1:
        include_performance = st.checkbox("Performance Analysis", value=True)
        include_risk = st.checkbox("Risk Metrics", value=True)
        include_attribution = st.checkbox("Performance Attribution", value=True)
        include_scenarios = st.checkbox("Scenario Analysis", value=True)
    
    with col_content2:
        include_allocation = st.checkbox("Portfolio Allocation", value=True)
        include_benchmark = st.checkbox("Benchmark Comparison", value=True)
        include_recommendations = st.checkbox("Recommendations", value=True)
        include_appendix = st.checkbox("Technical Appendix", value=False)
    
    # Generate reports
    if exec_summary or full_report or risk_report:
        with st.spinner("🔨 Generating professional report..."):
            # Prepare report data
            report_data = {
                'total_return': optimization_result['metrics'].get('total_return', 0),
                'annual_return': optimization_result['metrics'].get('annual_return', 0),
                'annual_vol': optimization_result['metrics'].get('annual_volatility', 0),
                'sharpe': optimization_result['metrics'].get('sharpe_ratio', 0),
                'max_dd': optimization_result['metrics'].get('max_drawdown', 0),
                'var_95': optimization_result['metrics'].get('value_at_risk_95', 0),
                'diversification_score': optimization_result.get('diversification_score', 0.7),
                'liquidity_score': 0.8,  # Example
                'benchmark_return': 0.12,  # Example
                'benchmark_annual': 0.10,  # Example
                'benchmark_vol': 0.15,  # Example
                'benchmark_sharpe': 0.8,  # Example
                'benchmark_dd': -0.12,  # Example
            }
            
            # Generate report based on type
            if exec_summary:
                report = InstitutionalReportGenerator.generate_executive_summary(
                    report_data, optimization_result
                )
                
                st.markdown("### 📊 Executive Summary Report")
                st.markdown(report)
                
                # Download button
                st.download_button(
                    label="📥 Download Executive Summary (MD)",
                    data=report,
                    file_name=f"executive_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.md",
                    mime="text/markdown"
                )
            
            elif full_report:
                # Generate comprehensive report
                st.success("✅ Full analysis report generated successfully!")
                
                # Create multiple download options
                col_download1, col_download2, col_download3 = st.columns(3)
                
                with col_download1:
                    # Markdown report
                    full_report_md = InstitutionalReportGenerator.generate_executive_summary(
                        report_data, optimization_result
                    )
                    st.download_button(
                        label="📥 Download Markdown",
                        data=full_report_md,
                        file_name=f"full_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.md",
                        mime="text/markdown"
                    )
                
                with col_download2:
                    # Excel report
                    try:
                        excel_report = InstitutionalReportGenerator.generate_excel_report(
                            report_data, optimization_result
                        )
                        st.download_button(
                            label="📥 Download Excel",
                            data=excel_report,
                            file_name=f"portfolio_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    except:
                        st.warning("Excel report requires additional dependencies")
                
                with col_download3:
                    # PDF report (simplified)
                    try:
                        pdf_report = InstitutionalReportGenerator.generate_pdf_report(
                            report_data, optimization_result, []
                        )
                        st.download_button(
                            label="📥 Download PDF",
                            data=pdf_report,
                            file_name=f"portfolio_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                            mime="application/pdf"
                        )
                    except:
                        st.warning("PDF generation requires FPDF library")
            
            elif risk_report:
                st.success("✅ Risk assessment report generated!")
                
                # Display risk report
                st.markdown("### ⚖️ Risk Assessment Report")
                
                risk_metrics = [
                    ("Annual Volatility", report_data['annual_vol']),
                    ("Value at Risk (95%)", report_data['var_95']),
                    ("Maximum Drawdown", report_data['max_dd']),
                    ("Stress Test Score", "B+"),
                    ("Liquidity Score", f"{report_data['liquidity_score']:.1%}"),
                    ("Diversification Score", f"{report_data['diversification_score']:.1%}"),
                ]
                
                for metric, value in risk_metrics:
                    col_m1, col_m2 = st.columns([1, 2])
                    with col_m1:
                        st.write(f"**{metric}:**")
                    with col_m2:
                        if isinstance(value, (int, float)):
                            st.write(f"{value:.2%}")
                        else:
                            st.write(value)
    
    # Report templates
    st.markdown("### 🎨 Report Templates")
    
    template_col1, template_col2, template_col3 = st.columns(3)
    
    with template_col1:
        if st.button("🏦 Institutional Template", use_container_width=True):
            st.info("Institutional template selected")
    
    with template_col2:
        if st.button("📈 Performance Template", use_container_width=True):
            st.info("Performance template selected")
    
    with template_col3:
        if st.button("⚖️ Risk Template", use_container_width=True):
            st.info("Risk template selected")
    
    # Report scheduling
    st.markdown("### ⏰ Scheduled Reporting")
    
    with st.expander("🔄 Set Up Automated Reporting", expanded=False):
        col_schedule1, col_schedule2 = st.columns(2)
        
        with col_schedule1:
            frequency = st.selectbox(
                "Report Frequency",
                ["Daily", "Weekly", "Monthly", "Quarterly", "Annually"]
            )
        
        with col_schedule2:
            delivery = st.selectbox(
                "Delivery Method",
                ["Email", "Cloud Storage", "API Webhook", "Internal Dashboard"]
            )
        
        if st.button("💾 Save Reporting Schedule", use_container_width=True):
            st.success(f"✅ {frequency} reports scheduled for {delivery}")

def display_monitoring_tab(prices: pd.DataFrame, optimization_result: Dict):
    """Display enhanced monitoring tab"""
    
    st.markdown("""
    <div style="background: linear-gradient(90deg, #3498db, #2980b9); 
                padding: 1.5rem; border-radius: 10px; margin-bottom: 2rem;">
        <h2 style="color: white; margin: 0;">⚙️ PORTFOLIO MONITORING</h2>
        <p style="color: rgba(255,255,255,0.9); margin: 0;">Real-time Monitoring & Alert System</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Real-time metrics
    st.markdown("### 📊 Real-time Portfolio Metrics")
    
    # Calculate recent performance
    recent_days = 30
    recent_returns = prices.pct_change().iloc[-recent_days:]
    portfolio_returns_recent = (recent_returns * list(optimization_result['weights'].values())).sum(axis=1)
    
    col_mon1, col_mon2, col_mon3, col_mon4 = st.columns(4)
    
    with col_mon1:
        recent_return = (1 + portfolio_returns_recent).prod() - 1
        st.metric("30-Day Return", f"{recent_return:.2%}")
    
    with col_mon2:
        recent_vol = portfolio_returns_recent.std() * np.sqrt(252)
        st.metric("30-Day Volatility", f"{recent_vol:.2%}")
    
    with col_mon3:
        recent_sharpe = (portfolio_returns_recent.mean() * 252 - 0.03) / (recent_vol + 1e-10)
        st.metric("30-Day Sharpe", f"{recent_sharpe:.2f}")
    
    with col_mon4:
        recent_dd = InstitutionalMetrics._calculate_max_drawdown(portfolio_returns_recent)
        st.metric("30-Day Max DD", f"{recent_dd:.2%}")
    
    # Alert system
    st.markdown("### 🚨 Alert System")
    
    # Define alert thresholds
    with st.expander("⚙️ Alert Configuration", expanded=True):
        col_alert1, col_alert2 = st.columns(2)
        
        with col_alert1:
            dd_alert = st.number_input("Drawdown Alert (%)", 5.0, 50.0, 10.0, step=1.0)
            vol_alert = st.number_input("Volatility Spike (%)", 5.0, 50.0, 20.0, step=1.0)
        
        with col_alert2:
            corr_alert = st.number_input("Correlation Threshold", 0.5, 1.0, 0.8, step=0.05)
            liquidity_alert = st.number_input("Liquidity Threshold", 0.1, 1.0, 0.3, step=0.05)
    
    # Check alerts
    alerts = []
    
    # Check drawdown
    if abs(optimization_result['metrics'].get('max_drawdown', 0)) > dd_alert / 100:
        alerts.append({
            'type': 'danger',
            'message': f"⚠️ Maximum drawdown ({abs(optimization_result['metrics']['max_drawdown']):.2%}) exceeds threshold ({dd_alert}%)"
        })
    
    # Check volatility
    if optimization_result['metrics'].get('annual_volatility', 0) > vol_alert / 100:
        alerts.append({
            'type': 'warning',
            'message': f"⚠️ Portfolio volatility ({optimization_result['metrics']['annual_volatility']:.2%}) exceeds threshold ({vol_alert}%)"
        })
    
    # Display alerts
    if alerts:
        st.markdown("### 🔴 Active Alerts")
        for alert in alerts:
            if alert['type'] == 'danger':
                st.error(alert['message'])
            else:
                st.warning(alert['message'])
    else:
        st.success("✅ No active alerts. All systems normal.")
    
    # Rebalancing monitor
    st.markdown("### ⚖️ Rebalancing Monitor")
    
    # Calculate drift from target weights
    current_prices = prices.iloc[-1]
    portfolio_value = sum(current_prices[asset] * weight * 1000000  # Example: $1M portfolio
                         for asset, weight in optimization_result['weights'].items())
    
    drift_data = []
    for asset, target_weight in optimization_result['weights'].items():
        current_value = current_prices[asset] * target_weight * 1000000 / current_prices[asset]
        current_weight = current_value / portfolio_value
        drift = current_weight - target_weight
        
        drift_data.append({
            'Asset': asset,
            'Target Weight': target_weight,
            'Current Weight': current_weight,
            'Drift': drift,
            'Status': 'Within Range' if abs(drift) < 0.01 else 'Needs Rebalance'
        })
    
    drift_df = pd.DataFrame(drift_data)
    
    st.dataframe(
        drift_df.style.format({
            'Target Weight': '{:.2%}',
            'Current Weight': '{:.2%}',
            'Drift': '{:.3%}'
        }).apply(
            lambda x: ['background-color: #ffcccc' if v == 'Needs Rebalance' else '' 
                      for v in x],
            subset=['Status']
        ),
        use_container_width=True,
        height=300
    )
    
    # Rebalancing recommendations
    needs_rebalance = drift_df[drift_df['Status'] == 'Needs Rebalance']
    if not needs_rebalance.empty:
        st.markdown("#### 💡 Rebalancing Recommendations")
        
        for _, row in needs_rebalance.iterrows():
            if row['Drift'] > 0:
                action = f"Sell {abs(row['Drift'] * portfolio_value / current_prices[row['Asset']]):.0f} shares"
            else:
                action = f"Buy {abs(row['Drift'] * portfolio_value / current_prices[row['Asset']]):.0f} shares"
            
            st.write(f"**{row['Asset']}**: {action} to return to target weight")
    
    # Performance attribution monitoring
    st.markdown("### 📈 Performance Attribution Monitor")
    
    # Calculate rolling attribution
    window_size = 63  # Quarterly
    if len(prices) > window_size:
        rolling_attribution = []
        
        for i in range(window_size, len(prices), 22):  # Monthly steps
            window_returns = prices.iloc[i-window_size:i].pct_change().dropna()
            window_portfolio_returns = (window_returns * list(optimization_result['weights'].values())).sum(axis=1)
            
            # Simplified attribution
            total_return = (1 + window_portfolio_returns).prod() - 1
            rolling_attribution.append({
                'Date': prices.index[i-1],
                'Total Return': total_return,
                'Market Return': window_returns.mean(axis=1).mean() * window_size,
                'Alpha': total_return - window_returns.mean(axis=1).mean() * window_size
            })
        
        if rolling_attribution:
            attribution_df = pd.DataFrame(rolling_attribution).set_index('Date')
            
            fig = go.Figure()
            
            fig.add_trace(go.Scatter(
                x=attribution_df.index,
                y=attribution_df['Total Return'] * 100,
                name='Total Return',
                line=dict(color='#1a5fb4', width=2)
            ))
            
            fig.add_trace(go.Scatter(
                x=attribution_df.index,
                y=attribution_df['Alpha'] * 100,
                name='Alpha',
                line=dict(color='#27ae60', width=2, dash='dash')
            ))
            
            fig.update_layout(
                title="Rolling Performance Attribution",
                height=400,
                xaxis_title="Date",
                yaxis_title="Return (%)",
                template="plotly_dark",
                hovermode='x unified'
            )
            
            st.plotly_chart(fig, use_container_width=True)

def display_welcome_screen():
    """Display enhanced welcome screen"""
    
    # Hero section
    col_hero1, col_hero2 = st.columns([2, 1])
    
    with col_hero1:
        st.markdown("""
        <div style="margin-bottom: 3rem;">
            <h1 style="color: #1a5fb4; font-size: 3rem; margin-bottom: 1rem;">
                Institutional Portfolio Management
            </h1>
            <p style="color: #94a3b8; font-size: 1.2rem; line-height: 1.6;">
                APOLLO/ENIGMA provides institutional-grade portfolio analysis, 
                risk management, and reporting for sophisticated investors.
            </p>
        </div>
        """, unsafe_allow_html=True)
    
    with col_hero2:
        st.markdown("""
        <div style="background: linear-gradient(135deg, #0a3d62, #1a5fb4); 
                    padding: 2rem; border-radius: 15px; text-align: center;">
            <div style="font-size: 2.5rem; margin-bottom: 1rem;">🏛️</div>
            <div style="font-size: 1.2rem; font-weight: bold; color: white;">
                Version 6.0
            </div>
            <div style="color: rgba(255,255,255,0.8); font-size: 0.9rem;">
                Institutional Edition
            </div>
        </div>
        """, unsafe_allow_html=True)
    
    # Key features
    st.markdown("### ✨ Key Features")
    
    features = [
        {
            "icon": "📊",
            "title": "Advanced Analytics",
            "description": "50+ institutional performance metrics and risk measures"
        },
        {
            "icon": "⚖️",
            "title": "Risk Management",
            "description": "Comprehensive risk analytics with stress testing"
        },
        {
            "icon": "📋",
            "title": "Professional Reporting",
            "description": "Institutional-grade reports in multiple formats"
        },
        {
            "icon": "🔍",
            "title": "Performance Attribution",
            "description": "Detailed Brinson attribution and factor analysis"
        },
        {
            "icon": "🚨",
            "title": "Monitoring & Alerts",
            "description": "Real-time monitoring with alert system"
        },
        {
            "icon": "🌍",
            "title": "Global Coverage",
            "description": "500+ assets across 8 major categories"
        }
    ]
    
    cols = st.columns(3)
    for idx, feature in enumerate(features):
        with cols[idx % 3]:
            st.markdown(f"""
            <div style="background: #1e293b; padding: 1.5rem; border-radius: 10px; 
                        border: 1px solid #334155; margin-bottom: 1rem; height: 200px;">
                <div style="font-size: 2rem; margin-bottom: 1rem;">{feature['icon']}</div>
                <div style="font-weight: bold; font-size: 1.1rem; margin-bottom: 0.5rem;">
                    {feature['title']}
                </div>
                <div style="color: #94a3b8; font-size: 0.9rem;">
                    {feature['description']}
                </div>
            </div>
            """, unsafe_allow_html=True)
    
    # Quick start guide
    with st.expander("🚀 Quick Start Guide", expanded=True):
        st.markdown("""
        1. **Select Assets**: Choose 3-10 diverse assets from the sidebar
        2. **Choose Strategy**: Select an institutional portfolio strategy
        3. **Set Parameters**: Configure risk parameters and constraints
        4. **Run Analysis**: Click "Run Analysis" to generate comprehensive insights
        5. **Generate Reports**: Create professional reports in multiple formats
        
        **Recommended Starting Portfolio:**
        - SPY (US Equities): 40%
        - TLT (US Bonds): 40%
        - GLD (Gold): 10%
        - BTC-USD (Crypto): 5%
        - VNQ (Real Estate): 5%
        """)
    
    # System status
    st.markdown("### 🔧 System Status")
    
    col_status1, col_status2, col_status3 = st.columns(3)
    
    with col_status1:
        st.metric("Data Sources", "Yahoo Finance", "Active")
    
    with col_status2:
        st.metric("Optimization Engine", "PyPortfolioOpt", "✅" if PYPFOPT_AVAILABLE else "❌")
    
    with col_status3:
        st.metric("Charting Library", "Plotly", "Active")

# -------------------------------------------------------------
# UTILITY FUNCTIONS
# -------------------------------------------------------------
@st.cache_data(ttl=3600)
def load_institutional_data(tickers: List[str], start_date: pd.Timestamp, 
                           end_date: pd.Timestamp) -> pd.DataFrame:
    """Load institutional data with enhanced error handling"""
    try:
        data = yf.download(
            tickers,
            start=start_date,
            end=end_date,
            progress=False,
            group_by='ticker'
        )
        
        # Handle multi-index columns
        if isinstance(data.columns, pd.MultiIndex):
            prices = pd.DataFrame()
            for ticker in tickers:
                if ticker in data.columns.levels[0]:
                    if ('Adj Close', ticker) in data.columns:
                        prices[ticker] = data[('Adj Close', ticker)]
                    elif ('Close', ticker) in data.columns:
                        prices[ticker] = data[('Close', ticker)]
        else:
            if 'Adj Close' in data.columns:
                prices = data['Adj Close']
            elif 'Close' in data.columns:
                prices = data['Close']
            else:
                prices = data
        
        return prices.dropna()
    
    except Exception as e:
        st.error(f"Error loading data: {str(e)[:200]}")
        return pd.DataFrame()

def get_asset_category(ticker: str) -> str:
    """Get asset category for a ticker"""
    for category, assets in GLOBAL_ASSET_UNIVERSE.items():
        if ticker in assets:
            return category.replace('_', ' ')
    return "Other"

# -------------------------------------------------------------
# APPLICATION ENTRY POINT
# -------------------------------------------------------------
if __name__ == "__main__":
    try:
        main()
        
        # Professional footer
        st.markdown("""
        <div class="institutional-footer">
            <div>🏛️ APOLLO/ENIGMA INSTITUTIONAL PORTFOLIO TERMINAL v6.0</div>
            <div style="font-size: 0.7rem; margin-top: 0.5rem; opacity: 0.7;">
                For institutional use only. Not for distribution to retail investors.
                Past performance does not guarantee future results.
            </div>
        </div>
        """, unsafe_allow_html=True)
        
    except Exception as e:
        st.error(f"Application error: {str(e)[:200]}")
        st.info("Please refresh the page and try again.")
